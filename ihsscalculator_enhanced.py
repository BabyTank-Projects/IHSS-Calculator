"""
Overtime Hours Calendar Generator (Enhanced Version)
- Modern UI with improved colors and spacing
- Calendar grid with per-day hour entries (accepts: "8", "8.5", "8:30")
- Work time calculator: input start time to calculate end times
- Weekly totals update live
- Auto-fill distributes authorized hours across selected workdays
- Export Excel "timesheet_YYYY_MM_<period>.xlsx" with professional formatting and colors
- Help button with feature explanations

DISCLAIMER:
This is a personal scheduling/calculation tool. Always bill ONLY hours actually worked,
and follow IHSS rules/laws. The max-weekly calculation here is a simple helper and may
not match every IHSS situation. Use official guidance when in doubt.
"""

import calendar
import csv
import math
import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from datetime import date, datetime, timedelta
from pathlib import Path


# -----------------------------
# Compatibility helper
# -----------------------------
def _trace_write(var, callback):
    """Attach a write-trace callback for Tkinter Variables across Python/Tk builds.
    Uses trace_add when available; falls back to trace for older Tkinter.
    Callback must accept *args.
    """
    try:
        return var.trace_add("write", callback)
    except AttributeError:
        # Older tkinter uses trace(mode, callback)
        return var.trace("w", callback)


def get_safe_output_directory():
    """
    Returns a safe directory for saving files that works both when running
    as a script and as a PyInstaller executable.
    
    Priority:
    1. User's Documents folder
    2. User's Desktop
    3. User's home directory
    4. Temp directory as last resort
    """
    try:
        # Try Documents folder first (most appropriate)
        if os.name == 'nt':  # Windows
            import winreg
            try:
                # Get Documents folder from Windows registry
                key = winreg.OpenKey(
                    winreg.HKEY_CURRENT_USER,
                    r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
                )
                documents_path = winreg.QueryValueEx(key, 'Personal')[0]
                winreg.CloseKey(key)
                if os.path.exists(documents_path) and os.access(documents_path, os.W_OK):
                    return documents_path
            except:
                pass
            
            # Fallback: standard Documents path
            documents = Path.home() / "Documents"
            if documents.exists() and os.access(documents, os.W_OK):
                return str(documents)
        else:  # macOS, Linux
            documents = Path.home() / "Documents"
            if documents.exists() and os.access(documents, os.W_OK):
                return str(documents)
    except:
        pass
    
    # Try Desktop
    try:
        desktop = Path.home() / "Desktop"
        if desktop.exists() and os.access(desktop, os.W_OK):
            return str(desktop)
    except:
        pass
    
    # Try home directory
    try:
        home = Path.home()
        if home.exists() and os.access(home, os.W_OK):
            return str(home)
    except:
        pass
    
    # Last resort: temp directory
    import tempfile
    return tempfile.gettempdir()


# -----------------------------
# Helpers: parsing + formatting
# -----------------------------
def parse_duration_to_minutes(text: str) -> int:
    """
    Accepts:
      - "" -> 0
      - "8" -> 480
      - "8.5" -> 510
      - "8:30" -> 510
      - "0:45" -> 45
    Returns minutes (int).
    """
    s = (text or "").strip()
    if not s:
        return 0

    # H:MM format
    if ":" in s:
        parts = s.split(":")
        if len(parts) != 2:
            raise ValueError("Use H:MM (e.g., 7:30).")
        h = int(parts[0].strip())
        m = int(parts[1].strip())
        if m < 0 or m >= 60:
            raise ValueError("Minutes must be 0-59.")
        if h < 0:
            raise ValueError("Hours must be >= 0.")
        return h * 60 + m

    # Decimal hours
    try:
        hours = float(s)
    except ValueError:
        raise ValueError('Enter hours like "8", "8.5", or "8:30".')

    if hours < 0:
        raise ValueError("Hours must be >= 0.")
    return int(round(hours * 60))


def minutes_to_h_mm(total_minutes: int) -> str:
    if total_minutes < 0:
        total_minutes = 0
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h}:{m:02d}"


def parse_time(time_str: str) -> tuple:
    """Parse time string in format HH:MM or H:MM (12/24 hour)"""
    time_str = time_str.strip().upper()
    
    # Handle AM/PM
    is_pm = 'PM' in time_str
    is_am = 'AM' in time_str
    time_str = time_str.replace('AM', '').replace('PM', '').strip()
    
    parts = time_str.split(':')
    if len(parts) != 2:
        raise ValueError("Use format HH:MM (e.g., 9:00 or 14:30)")
    
    hours = int(parts[0])
    minutes = int(parts[1])
    
    # Convert to 24-hour format if AM/PM specified
    if is_pm and hours != 12:
        hours += 12
    elif is_am and hours == 12:
        hours = 0
    
    if hours < 0 or hours >= 24:
        raise ValueError("Hours must be 0-23")
    if minutes < 0 or minutes >= 60:
        raise ValueError("Minutes must be 0-59")
    
    return (hours, minutes)


def format_time(hours: int, minutes: int) -> str:
    """Format time as HH:MM AM/PM"""
    period = "AM" if hours < 12 else "PM"
    display_hour = hours % 12
    if display_hour == 0:
        display_hour = 12
    return f"{display_hour}:{minutes:02d} {period}"


def safe_int(text: str) -> int:
    s = (text or "").strip()
    if not s:
        return 0
    return int(s)


def clamp(n: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, n))


# -----------------------------
# Help Window
# -----------------------------
class HelpWindow:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Help - Feature Guide")
        self.window.geometry("700x600")
        self.window.configure(bg='#1e1e2e')
        
        # Main frame
        main_frame = ttk.Frame(self.window, padding=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title = ttk.Label(main_frame, text="📖 IHSS Calculator Help", 
                         font=('Segoe UI', 16, 'bold'),
                         foreground='#f5e0dc',
                         background='#1e1e2e')
        title.pack(pady=(0, 15))
        
        # Scrolled text for help content
        help_text = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, 
                                              font=('Segoe UI', 10),
                                              bg='#313244', 
                                              fg='#cdd6f4',
                                              insertbackground='#cdd6f4',
                                              padx=15, pady=15)
        help_text.pack(fill="both", expand=True)
        
        # Help content
        content = """
MONTHLY AUTHORIZED HOURS
Enter the total hours and minutes you're authorized to work per month. This is used to calculate how many hours to distribute across your work schedule.

EXEMPTIONS
These checkboxes are helpers for calculating weekly maximum hours:
• Exemption 1 & 2: Check if you qualify for 90 hours/week exemptions
• Override: Manually set a custom weekly maximum if you have special authorization

WORKDAYS SELECTION
Check the days you typically work (Su=Sunday, M=Monday, Tu=Tuesday, W=Wednesday, Th=Thursday, F=Friday, Sa=Saturday). If you leave all unchecked, the system assumes you work every day. This affects auto-fill distribution.

USE WHOLE HOURS
When checked, the auto-fill feature will try to assign whole hours (8:00, 7:00, etc.) instead of fractional hours (7:30, 6:45), with any leftover minutes concentrated on one day for cleaner scheduling.

PERIOD SELECTION
Choose between:
• Full Month: Entire month's hours
• Pay Period 1 (1-15): First half of month (60% of monthly hours)
• Pay Period 2 (16-end): Second half of month (40% of monthly hours)

CALENDAR GRID
• Enter daily hours in each cell
• Accepts formats: "8" (8 hours), "8.5" (8½ hours), "8:30" (8 hours 30 min)
• Weekly totals display on the right
• Cells are color-coded:
  - Green: Below/at weekly maximum
  - Yellow/Orange: Approaching maximum
  - Red: Exceeds weekly maximum

WORK TIME CALCULATOR
Enter your start time to see when you'll finish each day:
• Enter time as HH:MM (e.g., "9:00" or "14:30")
• Can use AM/PM format (e.g., "9:00 AM" or "2:30 PM")
• Click "Calculate End Times" to see finish times for all scheduled days
• Results appear in the calendar below each day's hours

BUTTONS
• Clear Calendar: Removes all entered hours
• Auto-Fill: Automatically distributes your authorized hours across selected workdays
• Export Excel: Saves your timesheet as a beautifully formatted Excel file (.xlsx) with colors and professional styling. If you use the Work Time Calculator, start and end times are automatically included!
• Screenshot: Takes a screenshot of the calendar only (useful for visual reference when filling timesheets)

TIPS
• Always verify your schedule matches actual hours worked
• Use Export Excel to keep professional records of your timesheets with beautiful formatting
• Use Screenshot to save a visual copy of your calendar for reference
• The weekly maximum calculations are helpers - follow official IHSS guidelines
• You can mix manual entry with auto-fill by clearing specific days after auto-filling
        """
        
        help_text.insert('1.0', content.strip())
        help_text.configure(state='disabled')  # Make read-only
        
        # Close button
        close_btn = ttk.Button(main_frame, text="Close", command=self.window.destroy)
        close_btn.pack(pady=(15, 0))


# -----------------------------
# Main App
# -----------------------------
class OvertimeCalendarApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("IHSS Hours Calendar - Enhanced")
        self.root.geometry("1350x820")
        
        # Configure dark theme colors
        self.colors = {
            'bg_dark': '#1e1e2e',          # Dark background
            'bg_darker': '#181825',        # Darker sections
            'bg_card': '#313244',          # Card backgrounds
            'bg_input': '#45475a',         # Input fields
            'primary': '#89b4fa',          # Blue accent
            'primary_hover': '#b4befe',    # Lighter blue
            'success': '#a6e3a1',          # Green
            'warning': '#f9e2af',          # Yellow/Orange
            'danger': '#f38ba8',           # Red/Pink
            'text': '#cdd6f4',             # Main text
            'text_bright': '#f5e0dc',      # Bright text
            'text_muted': '#9399b2',       # Muted text
            'border': '#585b70'            # Borders
        }
        
        self.root.configure(bg=self.colors['bg_dark'])

        # State vars
        today = date.today()
        self.month_var = tk.IntVar(value=today.month)
        self.year_var = tk.IntVar(value=today.year)
        self.period_var = tk.StringVar(value="Full Month")

        self.month_hours_var = tk.StringVar(value="0")
        self.month_minutes_var = tk.StringVar(value="0")

        self.ex1_var = tk.BooleanVar(value=False)
        self.ex2_var = tk.BooleanVar(value=False)
        self.override_enabled_var = tk.BooleanVar(value=False)
        self.override_hours_var = tk.StringVar(value="0")
        self.override_minutes_var = tk.StringVar(value="0")

        self.use_whole_hours_var = tk.BooleanVar(value=True)
        
        # Work time calculator vars
        self.start_time_var = tk.StringVar(value="9:00 AM")
        self.show_end_times_var = tk.BooleanVar(value=False)

        # Workdays checkboxes (Sun..Sat)
        self.workday_vars = [tk.BooleanVar(value=False) for _ in range(7)]

        # Calendar storage
        self.day_vars = {}           # date -> StringVar (user entry)
        self.day_widgets = {}        # date -> Entry widget
        self.end_time_labels = {}    # date -> Label for end time
        self.week_total_labels = []

        # Apply modern style
        self._apply_modern_style()
        
        # Layout
        self._build_ui()
        self._render_calendar()

    def _apply_modern_style(self):
        """Apply dark theme styling to ttk widgets"""
        style = ttk.Style()
        
        # Use clam theme as base for better customization
        try:
            available = set(style.theme_names() or ())
            if 'clam' in available:
                style.theme_use('clam')
        except Exception:
            pass
        
        # Configure dark theme styles
        style.configure('.', 
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text'],
                       fieldbackground=self.colors['bg_input'],
                       bordercolor=self.colors['border'])
        
        style.configure('TFrame', 
                       background=self.colors['bg_dark'])
        
        style.configure('TLabelframe', 
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_bright'],
                       bordercolor=self.colors['border'],
                       relief='solid',
                       borderwidth=1)
        
        style.configure('TLabelframe.Label',
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_bright'],
                       font=('Segoe UI', 10, 'bold'))
        
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 13, 'bold'),
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_bright'])
        
        style.configure('Header.TLabel', 
                       font=('Segoe UI', 9, 'bold'),
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_bright'])
        
        style.configure('Muted.TLabel',
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_muted'],
                       font=('Segoe UI', 9))
        
        style.configure('TLabel',
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text'])
        
        style.configure('TButton',
                       background=self.colors['primary'],
                       foreground=self.colors['bg_dark'],
                       bordercolor=self.colors['primary'],
                       focuscolor=self.colors['primary_hover'],
                       font=('Segoe UI', 9, 'bold'),
                       padding=(10, 5))
        
        style.map('TButton',
                 background=[('active', self.colors['primary_hover']),
                           ('pressed', self.colors['primary'])],
                 foreground=[('active', self.colors['bg_dark'])])
        
        style.configure('Primary.TButton',
                       background=self.colors['primary'],
                       foreground=self.colors['bg_dark'],
                       font=('Segoe UI', 9, 'bold'),
                       padding=(10, 5))
        
        style.map('Primary.TButton',
                 background=[('active', self.colors['primary_hover'])],
                 foreground=[('active', self.colors['bg_dark'])])
        
        style.configure('TCheckbutton',
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text'])
        
        style.map('TCheckbutton',
                 background=[('active', self.colors['bg_dark'])])
        
        style.configure('TEntry',
                       fieldbackground=self.colors['bg_input'],
                       foreground=self.colors['text_bright'],
                       bordercolor=self.colors['border'],
                       insertcolor=self.colors['text_bright'])
        
        style.configure('TCombobox',
                       fieldbackground=self.colors['bg_input'],
                       background=self.colors['bg_input'],
                       foreground=self.colors['text_bright'],
                       bordercolor=self.colors['border'],
                       arrowcolor=self.colors['text'],
                       selectbackground=self.colors['primary'],
                       selectforeground=self.colors['bg_dark'])
        
        style.map('TCombobox',
                 fieldbackground=[('readonly', self.colors['bg_input'])],
                 selectbackground=[('readonly', self.colors['bg_input'])])

    def _build_ui(self):
        today = date.today()
        
        # Main container with padding
        outer = ttk.Frame(self.root, padding=10)
        outer.grid(row=0, column=0, sticky="nsew")
        
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Header with title and help button
        header_frame = ttk.Frame(outer)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        
        title_label = ttk.Label(header_frame, text="⏰ IHSS Hours Calendar Generator", 
                               style='Title.TLabel')
        title_label.pack(side="left")
        
        help_btn = ttk.Button(header_frame, text="❓ Help", command=self.show_help,
                             style='Primary.TButton')
        help_btn.pack(side="right")

        # LEFT COLUMN - Settings
        left_column = ttk.Frame(outer, width=300)
        left_column.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        left_column.grid_propagate(False)  # Prevent column from shrinking
        
        # Settings section
        settings = ttk.Labelframe(left_column, text="⚙️ Settings", padding=8)
        settings.pack(fill="both", expand=False, pady=(0, 8))

        # Monthly hours
        ttk.Label(settings, text="Monthly Hours:", style='Header.TLabel').grid(
            row=0, column=0, sticky="w", pady=(0, 3))
        
        hours_frame = ttk.Frame(settings)
        hours_frame.grid(row=1, column=0, sticky="w", pady=(0, 8))
        ttk.Entry(hours_frame, textvariable=self.month_hours_var, 
                 width=6, font=('Segoe UI', 9)).pack(side="left", padx=(0, 2))
        ttk.Label(hours_frame, text="h").pack(side="left", padx=(0, 8))
        ttk.Entry(hours_frame, textvariable=self.month_minutes_var, 
                 width=6, font=('Segoe UI', 9)).pack(side="left", padx=(0, 2))
        ttk.Label(hours_frame, text="m").pack(side="left")

        # Exemptions section
        ttk.Label(settings, text="Exemptions:", style='Header.TLabel').grid(
            row=2, column=0, sticky="w", pady=(5, 3))
        
        ttk.Checkbutton(settings, text="Live-In Family Care Provider",
                       variable=self.ex1_var, 
                       command=self._update_max_weekly_label).grid(
                           row=3, column=0, sticky="w")
        
        ttk.Checkbutton(settings, text="Extraordinary Circumstances",
                       variable=self.ex2_var, 
                       command=self._update_max_weekly_label).grid(
                           row=4, column=0, sticky="w")

        # Override
        ttk.Checkbutton(settings, text="Override weekly max:",
                       variable=self.override_enabled_var, 
                       command=self._update_max_weekly_label).grid(
                           row=5, column=0, sticky="w", pady=(3, 2))
        
        override_frame = ttk.Frame(settings)
        override_frame.grid(row=6, column=0, sticky="w", pady=(0, 5))
        ttk.Entry(override_frame, textvariable=self.override_hours_var, 
                 width=5, font=('Segoe UI', 9)).pack(side="left", padx=(0, 2))
        ttk.Label(override_frame, text="h").pack(side="left", padx=(0, 5))
        ttk.Entry(override_frame, textvariable=self.override_minutes_var, 
                 width=5, font=('Segoe UI', 9)).pack(side="left", padx=(0, 2))
        ttk.Label(override_frame, text="m/wk").pack(side="left")

        # Max weekly label
        self.max_weekly_label = ttk.Label(settings, text="Max Weekly: --", 
                                         style='Muted.TLabel', wraplength=200)
        self.max_weekly_label.grid(row=7, column=0, sticky="w", pady=(0, 5))

        # Workdays
        ttk.Label(settings, text="Work Days:", style='Header.TLabel').grid(
            row=8, column=0, sticky="w", pady=(3, 2))
        
        ttk.Label(settings, text="(uncheck all = every day)", 
                 font=('Segoe UI', 8), foreground=self.colors['text_muted']).grid(
                     row=9, column=0, sticky="w", pady=(0, 3))

        days_frame = ttk.Frame(settings)
        days_frame.grid(row=10, column=0, sticky="w", pady=(0, 5))
        
        days = ["Su", "M", "Tu", "W", "Th", "F", "Sa"]
        day_full = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        for i, (d, full) in enumerate(zip(days, day_full)):
            btn = ttk.Checkbutton(days_frame, text=d, width=3,
                                 variable=self.workday_vars[i])
            btn.pack(side="left", padx=2)
            # Create tooltip effect (not visible but helps code clarity)
            # Tooltip: btn represents day_full[i]

        # Options
        ttk.Checkbutton(settings, text="Use whole hours",
                       variable=self.use_whole_hours_var).grid(
                           row=11, column=0, sticky="w", pady=(3, 5))

        # Period/Month/Year selection
        ttk.Label(settings, text="Period:", style='Header.TLabel').grid(
            row=12, column=0, sticky="w", pady=(3, 3))
        
        period_combo = ttk.Combobox(settings, textvariable=self.period_var, 
                                   state="readonly", width=22,
                                   values=["Full Month", "Pay Period 1 (1-15)", 
                                          "Pay Period 2 (16-end)"])
        period_combo.grid(row=13, column=0, sticky="w", pady=(0, 5))
        period_combo.bind("<<ComboboxSelected>>", lambda _: self._render_calendar())
        
        ttk.Label(settings, text="Month:", style='Header.TLabel').grid(
            row=14, column=0, sticky="w", pady=(3, 3))
        month_combo = ttk.Combobox(settings, textvariable=self.month_var, 
                                  state="readonly", width=22,
                                  values=list(range(1, 13)))
        month_combo.grid(row=15, column=0, sticky="w", pady=(0, 5))
        month_combo.bind("<<ComboboxSelected>>", lambda _: self._render_calendar())
        
        ttk.Label(settings, text="Year:", style='Header.TLabel').grid(
            row=16, column=0, sticky="w", pady=(3, 3))
        year_combo = ttk.Combobox(settings, textvariable=self.year_var, 
                                 state="readonly", width=22,
                                 values=list(range(today.year - 3, 2127)))
        year_combo.grid(row=17, column=0, sticky="w", pady=(0, 5))
        year_combo.bind("<<ComboboxSelected>>", lambda _: self._render_calendar())

        # Work Time Calculator Section
        time_calc_frame = ttk.Labelframe(left_column, text="🕐 Work Time Calculator", padding=8)
        time_calc_frame.pack(fill="both", expand=False)
        
        ttk.Label(time_calc_frame, text="Start Time:", 
                 style='Header.TLabel').pack(anchor="w", pady=(0, 2))
        
        start_entry = ttk.Entry(time_calc_frame, textvariable=self.start_time_var, 
                               width=25, font=('Segoe UI', 9))
        start_entry.pack(anchor="w", fill="x", pady=(0, 2))
        
        ttk.Label(time_calc_frame, text="(e.g., 9:00 AM)", 
                 font=('Segoe UI', 8), foreground=self.colors['text_muted']).pack(
                     anchor="w", pady=(0, 6))
        
        ttk.Button(time_calc_frame, text="📅 Calculate", 
                  command=self.calculate_end_times,
                  style='Primary.TButton').pack(fill="x", pady=(0, 6))
        
        ttk.Checkbutton(time_calc_frame, text="Show in calendar", 
                       variable=self.show_end_times_var).pack(anchor="w", pady=(0, 2))

        # RIGHT COLUMN - Calendar section
        calendar_frame = ttk.Labelframe(outer, text="📆 Calendar", padding=10)
        calendar_frame.grid(row=1, column=1, sticky="nsew")
        
        outer.grid_rowconfigure(1, weight=1)
        outer.grid_columnconfigure(1, weight=1)
        
        # Calendar inner frame (no scrollbar)
        self.calendar_inner = tk.Frame(calendar_frame, bg=self.colors['bg_darker'])
        self.calendar_inner.pack(fill="both", expand=True)
        
        # Buttons section - below calendar
        btn_frame = ttk.Frame(outer)
        btn_frame.grid(row=2, column=1, sticky="ew", pady=(10, 0))
        
        ttk.Button(btn_frame, text="🗑️ Clear Calendar", 
                  command=self.clear_calendar,
                  style='Primary.TButton').pack(side="left", padx=(0, 10))
        
        ttk.Button(btn_frame, text="✨ Auto-Fill", 
                  command=self.autofill,
                  style='Primary.TButton').pack(side="left", padx=(0, 10))
        
        ttk.Button(btn_frame, text="💾 Export Excel", 
                  command=self.export_csv,
                  style='Primary.TButton').pack(side="left", padx=(0, 10))
        
        ttk.Button(btn_frame, text="📸 Screenshot", 
                  command=self.take_screenshot,
                  style='Primary.TButton').pack(side="left")

        # Attach trace callbacks
        _trace_write(self.month_hours_var, lambda *_: self._update_max_weekly_label())
        _trace_write(self.month_minutes_var, lambda *_: self._update_max_weekly_label())
        _trace_write(self.override_hours_var, lambda *_: self._update_max_weekly_label())
        _trace_write(self.override_minutes_var, lambda *_: self._update_max_weekly_label())

        self._update_max_weekly_label()

    def show_help(self):
        """Show help window"""
        HelpWindow(self.root)

    def calculate_end_times(self):
        """Calculate and display end times based on start time and daily hours"""
        try:
            start_hours, start_minutes = parse_time(self.start_time_var.get())
        except ValueError as e:
            messagebox.showerror("Invalid Time", str(e))
            return
        
        self.show_end_times_var.set(True)
        
        # Calculate for each day with hours
        results = []
        for dt, var in self.day_vars.items():
            text = var.get().strip()
            if not text:
                if dt in self.end_time_labels:
                    self.end_time_labels[dt].config(text="")
                continue
            
            try:
                work_minutes = parse_duration_to_minutes(text)
                if work_minutes == 0:
                    if dt in self.end_time_labels:
                        self.end_time_labels[dt].config(text="")
                    continue
                
                # Calculate end time
                total_minutes = start_hours * 60 + start_minutes + work_minutes
                end_hours = (total_minutes // 60) % 24
                end_minutes = total_minutes % 60
                
                end_time_str = format_time(end_hours, end_minutes)
                
                # Update label if it exists
                if dt in self.end_time_labels:
                    self.end_time_labels[dt].config(
                        text=f"→ {end_time_str}",
                        foreground=self.colors['success']
                    )
                
                results.append(f"{dt.strftime('%a %m/%d')}: {format_time(start_hours, start_minutes)} → {end_time_str}")
                
            except ValueError:
                if dt in self.end_time_labels:
                    self.end_time_labels[dt].config(text="")
                continue
        
        if results:
            messagebox.showinfo("End Times Calculated", 
                              f"Work schedule calculated!\n\n" + "\n".join(results[:10]) + 
                              ("\n..." if len(results) > 10 else ""))

    def _render_calendar(self):
        """Render the calendar grid for the selected period"""
        # Clear existing
        for widget in self.calendar_inner.winfo_children():
            widget.destroy()
        
        self.day_vars.clear()
        self.day_widgets.clear()
        self.end_time_labels.clear()
        self.week_total_labels.clear()

        y = int(self.year_var.get())
        m = int(self.month_var.get())
        
        # Header
        month_name = calendar.month_name[m]
        header = tk.Label(self.calendar_inner, 
                         text=f"{month_name} {y}", 
                         font=('Segoe UI', 16, 'bold'),
                         bg=self.colors['bg_darker'],
                         fg=self.colors['text_bright'])
        header.grid(row=0, column=0, columnspan=8, pady=(0, 10))

        # Day headers
        day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
        for col, name in enumerate(day_names):
            lbl = tk.Label(self.calendar_inner, text=name, 
                          font=('Segoe UI', 11, 'bold'),
                          bg=self.colors['bg_darker'],
                          fg=self.colors['text_bright'],
                          width=10)
            lbl.grid(row=1, column=col, padx=3, pady=5)
            self.calendar_inner.grid_columnconfigure(col, weight=1)
        
        tk.Label(self.calendar_inner, text="Weekly Total", 
                font=('Segoe UI', 11, 'bold'),
                bg=self.colors['bg_darker'],
                fg=self.colors['text_bright']).grid(row=1, column=7, padx=5)

        # Get dates in selected period
        dates = self._dates_in_selected_period()
        if not dates:
            return

        # Set calendar to start weeks on Sunday (6 = Sunday in Python's calendar)
        calendar.setfirstweekday(calendar.SUNDAY)
        
        # Calendar grid - now properly aligned with Sunday as first day
        cal = calendar.monthcalendar(y, m)
        row_offset = 2
        
        for week_idx, week in enumerate(cal):
            grid_row = row_offset + week_idx
            
            for day_of_week, day_num in enumerate(week):
                if day_num == 0:
                    continue
                
                dt = date(y, m, day_num)
                
                # Only create widgets for dates in selected period
                if dt not in dates:
                    tk.Label(self.calendar_inner, text=str(day_num),
                            bg=self.colors['bg_darker'],
                            fg=self.colors['text_muted']).grid(row=grid_row, column=day_of_week, padx=2, pady=2)
                    continue
                
                # Day cell frame
                cell_frame = tk.Frame(self.calendar_inner, 
                                     relief='solid', 
                                     borderwidth=1,
                                     bg=self.colors['bg_card'],
                                     highlightbackground=self.colors['border'],
                                     highlightthickness=1)
                cell_frame.grid(row=grid_row, column=day_of_week, padx=3, pady=3, sticky='nsew')
                
                # Make cells expand to fill space
                self.calendar_inner.grid_rowconfigure(grid_row, weight=1)
                self.calendar_inner.grid_columnconfigure(day_of_week, weight=1)
                
                # Day number
                day_label = tk.Label(cell_frame, text=str(day_num), 
                                    font=('Segoe UI', 11, 'bold'),
                                    bg=self.colors['bg_card'],
                                    fg=self.colors['text_bright'])
                day_label.pack(anchor='nw', padx=5, pady=3)
                
                # Hours entry
                var = tk.StringVar()
                self.day_vars[dt] = var
                
                entry = tk.Entry(cell_frame, textvariable=var, width=10, 
                               font=('Segoe UI', 12), justify='center',
                               bg=self.colors['bg_input'],
                               fg=self.colors['text_bright'],
                               insertbackground=self.colors['text_bright'],
                               relief='flat',
                               borderwidth=1,
                               highlightbackground=self.colors['border'],
                               highlightthickness=1)
                entry.pack(padx=5, pady=(0, 5), fill='x')
                self.day_widgets[dt] = entry
                
                # End time label (initially hidden)
                end_label = tk.Label(cell_frame, text="", 
                                    font=('Segoe UI', 10),
                                    bg=self.colors['bg_card'],
                                    fg=self.colors['success'])
                end_label.pack(padx=5, pady=(0, 5))
                self.end_time_labels[dt] = end_label
                
                # Trace for live updates
                _trace_write(var, lambda *_args, idx=week_idx: self._update_week_total(idx))
            
            # Weekly total label
            total_label = tk.Label(self.calendar_inner, text="0:00", 
                                  font=('Segoe UI', 12, 'bold'),
                                  bg=self.colors['bg_card'], 
                                  fg=self.colors['text_bright'],
                                  width=12, relief='solid', 
                                  borderwidth=1,
                                  highlightbackground=self.colors['border'],
                                  highlightthickness=1,
                                  pady=8)
            total_label.grid(row=grid_row, column=7, padx=5, pady=3, sticky='nsew')
            self.calendar_inner.grid_rowconfigure(grid_row, weight=1)
            self.week_total_labels.append(total_label)

        # Initial update
        for i in range(len(self.week_total_labels)):
            self._update_week_total(i)

    def _update_week_total(self, week_idx: int):
        """Update weekly total and color-code based on limits"""
        if week_idx >= len(self.week_total_labels):
            return

        y = int(self.year_var.get())
        m = int(self.month_var.get())
        cal = calendar.monthcalendar(y, m)
        
        if week_idx >= len(cal):
            return

        week = cal[week_idx]
        week_minutes = 0
        
        dates_in_period = set(self._dates_in_selected_period())

        for day_num in week:
            if day_num == 0:
                continue
            dt = date(y, m, day_num)
            
            if dt not in dates_in_period:
                continue
            
            if dt not in self.day_vars:
                continue

            text = self.day_vars[dt].get().strip()
            if not text:
                continue

            try:
                week_minutes += parse_duration_to_minutes(text)
            except ValueError:
                pass

        # Get max weekly
        max_weekly_minutes = self._max_weekly_minutes()
        
        # Update label
        label = self.week_total_labels[week_idx]
        label.config(text=minutes_to_h_mm(week_minutes))
        
        # Color code - simplified to just green or default
        if max_weekly_minutes > 0 and week_minutes > 0:
            # Just show green if there are hours
            label.config(bg=self.colors['success'], fg=self.colors['bg_dark'])
        else:
            # Default gray background if no hours
            label.config(bg=self.colors['bg_card'], fg=self.colors['text_bright'])

    def _max_weekly_minutes(self) -> int:
        """Calculate maximum weekly minutes based on exemptions"""
        if self.override_enabled_var.get():
            try:
                oh = safe_int(self.override_hours_var.get())
                om = safe_int(self.override_minutes_var.get())
                return oh * 60 + clamp(om, 0, 59)
            except:
                pass

        # Check exemptions
        if self.ex1_var.get() or self.ex2_var.get():
            return 90 * 60  # 90 hours

        # Default: 66 hours
        try:
            month_mins = self._month_authorized_minutes()
            # Approximate weekly from monthly
            return int(month_mins / 4.33)
        except:
            return 66 * 60

    def _update_max_weekly_label(self):
        """Update the max weekly hours label"""
        max_mins = self._max_weekly_minutes()
        self.max_weekly_label.config(
            text=f"Maximum Weekly Hours: {minutes_to_h_mm(max_mins)}"
        )
        
        # Refresh week totals
        for i in range(len(self.week_total_labels)):
            self._update_week_total(i)

    def _selected_workdays(self) -> set:
        """Return set of selected workday indices (0=Sun)"""
        return {i for i, var in enumerate(self.workday_vars) if var.get()}

    def _dates_in_selected_period(self):
        """Get list of dates in the selected period"""
        y = int(self.year_var.get())
        m = int(self.month_var.get())
        period = self.period_var.get()

        last_day = calendar.monthrange(y, m)[1]

        if period == "Pay Period 1 (1-15)":
            start_day, end_day = 1, min(15, last_day)
        elif period == "Pay Period 2 (16-end)":
            start_day, end_day = 16, last_day
        else:
            start_day, end_day = 1, last_day

        return [date(y, m, d) for d in range(start_day, end_day + 1)]

    def _month_authorized_minutes(self) -> int:
        """Get monthly authorized hours in minutes"""
        mh = safe_int(self.month_hours_var.get())
        mm = safe_int(self.month_minutes_var.get())
        return mh * 60 + clamp(mm, 0, 59)

    def _period_target_minutes(self) -> int:
        """Get target minutes for selected period"""
        month_minutes = self._month_authorized_minutes()
        period = self.period_var.get()

        if period == "Pay Period 1 (1-15)":
            return int(round(month_minutes * 0.60))
        if period == "Pay Period 2 (16-end)":
            return int(round(month_minutes * 0.40))
        return month_minutes

    def clear_calendar(self):
        """Clear all calendar entries"""
        for var in self.day_vars.values():
            var.set("")
        for label in self.end_time_labels.values():
            label.config(text="")
        for i in range(len(self.week_total_labels)):
            self._update_week_total(i)

    def autofill(self):
        """Auto-fill calendar with distributed hours"""
        try:
            month_minutes = self._month_authorized_minutes()
        except Exception:
            messagebox.showerror("Invalid Input", "Please enter valid monthly hours/minutes.")
            return

        if month_minutes <= 0:
            messagebox.showerror("Invalid Input", "Monthly authorized hours must be greater than 0.")
            return

        target_minutes = self._period_target_minutes()
        dates = self._dates_in_selected_period()
        selected_workdays = self._selected_workdays()

        def is_workday(dt: date) -> bool:
            # Python's weekday(): Monday=0, Sunday=6
            # We need: Sunday=0, Monday=1, ..., Saturday=6
            # Convert: (weekday + 1) % 7 gives us Sunday=0
            day_index = (dt.weekday() + 1) % 7
            if not selected_workdays:
                return True
            return day_index in selected_workdays

        work_dates = [dt for dt in dates if is_workday(dt)]

        if not work_dates:
            messagebox.showerror("No Workdays", "No days matched your workday selection.")
            return

        # Clear work dates
        for dt in work_dates:
            if dt in self.day_vars:
                self.day_vars[dt].set("")

        # Distribute
        n = len(work_dates)
        use_whole = self.use_whole_hours_var.get()

        if use_whole:
            import random
            total_hours = target_minutes // 60
            leftover_minutes = target_minutes % 60

            base_hours = total_hours // n
            extra_hours = total_hours % n

            allocations = [base_hours * 60] * n

            for i in range(extra_hours):
                allocations[i] += 60

            if leftover_minutes > 0:
                idx = random.randrange(n)
                allocations[idx] += leftover_minutes
        else:
            base = target_minutes // n
            remainder = target_minutes % n
            allocations = [base] * n
            for i in range(remainder):
                allocations[i] += 1

        # Write to calendar
        for dt, mins in zip(work_dates, allocations):
            if dt in self.day_vars:
                self.day_vars[dt].set(minutes_to_h_mm(mins))

        # Update totals
        for i in range(len(self.week_total_labels)):
            self._update_week_total(i)

        messagebox.showinfo(
            "Auto-Fill Complete",
            f"✅ Filled {len(work_dates)} day(s)\n"
            f"Total: {minutes_to_h_mm(target_minutes)}\n"
            f"Period: {self.period_var.get()}"
        )

    def take_screenshot(self):
        """Take a screenshot of only the calendar area"""
        from tkinter import simpledialog
        import datetime
        
        # Get default filename with current date
        default_name = f"calendar_screenshot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        # Ask user for filename
        filename = simpledialog.askstring(
            "Save Screenshot",
            "Enter filename for screenshot:",
            initialvalue=default_name
        )
        
        if not filename:
            return  # User cancelled
        
        # Ensure .png extension
        if not filename.lower().endswith('.png'):
            filename += '.png'
        
        try:
            # Get the calendar frame's position and size
            x = self.calendar_inner.winfo_rootx()
            y = self.calendar_inner.winfo_rooty()
            width = self.calendar_inner.winfo_width()
            height = self.calendar_inner.winfo_height()
            
            # Take screenshot using PIL
            try:
                from PIL import ImageGrab
            except ImportError:
                messagebox.showerror(
                    "Missing Library",
                    "Screenshot feature requires Pillow library.\n\n"
                    "Install with: pip install Pillow"
                )
                return
            
            # Capture the calendar area
            screenshot = ImageGrab.grab(bbox=(x, y, x + width, y + height))
            
            # Use safe output directory instead of cwd
            output_dir = get_safe_output_directory()
            filepath = os.path.join(output_dir, filename)
            screenshot.save(filepath, 'PNG')
            
            messagebox.showinfo(
                "Screenshot Saved",
                f"✅ Calendar screenshot saved:\n{filepath}"
            )
            
        except Exception as e:
            messagebox.showerror(
                "Screenshot Failed",
                f"Could not save screenshot:\n{str(e)}"
            )

    def export_csv(self):
        """Export timesheet to beautifully formatted Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "Excel export requires openpyxl library.\n\n"
                "Install with: pip install openpyxl"
            )
            return
            
        y = int(self.year_var.get())
        m = int(self.month_var.get())
        period = self.period_var.get()

        rows = []
        total_minutes = 0

        export_dates = self._dates_in_selected_period()

        # Parse work time calculator start time if available
        start_time_str = self.start_time_var.get().strip()
        using_work_time = bool(start_time_str)
        
        for dt in export_dates:
            if dt not in self.day_vars:
                continue
            text = self.day_vars[dt].get().strip()
            if not text:
                mins = 0
            else:
                try:
                    mins = parse_duration_to_minutes(text)
                except ValueError as e:
                    messagebox.showerror("Invalid Entry", f"{dt.isoformat()}: {e}")
                    return

            total_minutes += mins
            
            # Calculate end time if work time calculator is being used
            end_time_str = ""
            if using_work_time and mins > 0:
                try:
                    start_hours, start_minutes = parse_time(start_time_str)
                    end_hours, end_minutes = add_time(start_hours, start_minutes, mins)
                    end_time_str = format_time(end_hours, end_minutes)
                except:
                    end_time_str = ""
            
            rows.append({
                "date": dt.isoformat(),
                "day_name": dt.strftime("%A"),
                "hours_minutes": minutes_to_h_mm(mins),
                "decimal_hours": round(mins / 60.0, 2),
                "start_time": start_time_str if mins > 0 else "",
                "end_time": end_time_str,
                "minutes": mins
            })

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        
        # Define colors and styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")  # White text
        
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Medium blue
        title_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        
        total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
        total_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title section
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "⏰ IHSS HOURS TIMESHEET"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        
        # Info section
        ws['A3'] = "Year:"
        ws['B3'] = y
        ws['B3'].font = Font(bold=True)
        
        ws['A4'] = "Month:"
        ws['B4'] = calendar.month_name[m]
        ws['B4'].font = Font(bold=True)
        
        ws['A5'] = "Pay Period:"
        ws['B5'] = period
        ws['B5'].font = Font(bold=True)
        
        # Determine which columns to show
        start_row = 7
        col_headers = ["Date", "Day", "Hours (H:MM)", "Decimal Hours"]
        if using_work_time:
            col_headers.extend(["Start Time", "End Time"])
        
        # Column headers
        for col_idx, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Data rows
        current_row = start_row + 1
        for idx, r in enumerate(rows):
            # Apply alternating row colors
            row_fill = alt_row_fill if idx % 2 == 1 else None
            
            # Date
            cell = ws.cell(row=current_row, column=1, value=r["date"])
            cell.alignment = center_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            
            # Day name
            cell = ws.cell(row=current_row, column=2, value=r["day_name"])
            cell.alignment = left_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            
            # Hours (H:MM)
            cell = ws.cell(row=current_row, column=3, value=r["hours_minutes"])
            cell.alignment = center_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            
            # Decimal hours
            cell = ws.cell(row=current_row, column=4, value=r["decimal_hours"])
            cell.alignment = center_align
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            
            # Work time columns if applicable
            if using_work_time:
                # Start time
                cell = ws.cell(row=current_row, column=5, value=r["start_time"])
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                
                # End time
                cell = ws.cell(row=current_row, column=6, value=r["end_time"])
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
            
            current_row += 1
        
        # Total row
        current_row += 1
        total_col = 3 if not using_work_time else 4
        
        cell = ws.cell(row=current_row, column=1, value="TOTAL")
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        ws.merge_cells(f'A{current_row}:B{current_row}')
        
        cell = ws.cell(row=current_row, column=3, value=minutes_to_h_mm(total_minutes))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        cell = ws.cell(row=current_row, column=4, value=round(total_minutes / 60.0, 2))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        if using_work_time:
            for col in [5, 6]:
                cell = ws.cell(row=current_row, column=col, value="")
                cell.fill = total_fill
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        if using_work_time:
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
        
        # Filename - change to .xlsx
        safe_period = period.replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_").replace("/", "_")
        filename = f"timesheet_{y}_{m:02d}_{safe_period}.xlsx"
        
        # Use safe output directory
        output_dir = get_safe_output_directory()
        filepath = os.path.join(output_dir, filename)

        # Save Excel file
        try:
            wb.save(filepath)
        except OSError as e:
            messagebox.showerror("Export Failed", f"Could not write file:\n{filepath}\n\n{e}")
            return

        messagebox.showinfo(
            "Export Complete",
            f"✅ Saved Excel timesheet:\n{filepath}\n\n"
            f"Total Hours: {minutes_to_h_mm(total_minutes)}\n\n"
            f"📊 Professional formatting with colors applied!"
        )


def main():
    root = tk.Tk()
    app = OvertimeCalendarApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
